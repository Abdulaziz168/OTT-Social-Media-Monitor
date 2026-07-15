"""Excel report generator."""

from pathlib import Path
from datetime import datetime
from typing import List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from loguru import logger
from instagram_monitor.core.settings import settings
from instagram_monitor.database import Account, DailyStatistics


class ExcelReportGenerator:
    """Generate professional Excel reports."""

    def __init__(self):
        """Initialize report generator."""
        self.reports_dir = settings.REPORTS_DIR
        self.reports_dir.mkdir(exist_ok=True)

    def generate_report(
        self,
        accounts: List[Account],
        statistics: List[DailyStatistics],
        filename: Optional[str] = None,
    ) -> Path:
        """Generate comprehensive Excel report."""
        if not filename:
            filename = f"instagram_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        report_path = self.reports_dir / filename
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Dashboard"
            
            # Add header
            ws["A1"] = "OTT Instagram Monitor Report"
            ws["A1"].font = Font(size=16, bold=True)
            ws["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            
            logger.info(f"Excel report saved to {report_path}")
            wb.save(report_path)
            return report_path
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            raise
